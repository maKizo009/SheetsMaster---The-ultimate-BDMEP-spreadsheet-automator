import pandas as pd
import csv
import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import sys
import sheetMasters

janelaUsuario = tk.Tk()
larguraJanela = 500
alturaJanela = 300
janelaUsuario.title("SheetsMaster")
janelaUsuario.geometry(f"{larguraJanela}x{alturaJanela}")

barraProgresso = ttk.Progressbar(janelaUsuario, orient="horizontal", length=200, mode="determinate")
barraProgresso.pack()


def selecionandoCSV():

    global escolhaUsuario_CSV 
    escolhaUsuario_CSV = filedialog.askopenfilename(initialdir=os.path.expanduser("~/Downloads"), title="Selecione o CSV nessa pasta", filetypes=(("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")))

    processandoArquivo(escolhaUsuario_CSV)

def processandoArquivo(escolhaUsuario_CSV):

    if escolhaUsuario_CSV:

        print ("Arquivo selecionado: ", escolhaUsuario_CSV)

        with open (escolhaUsuario_CSV) as f_entrada:
            global leitor
            leitor = csv.reader(f_entrada, delimiter=";")

            global primeiraLinha 
            primeiraLinha = f_entrada.readline().rstrip()
            print(primeiraLinha)

            leitor = pd.DataFrame(leitor)

            leitor = leitor.iloc[11:]

            leitor = leitor.rename(columns={0: "Data", 1: "Max", 2: "Min", 3: "Precipitação"})
            leitor = leitor.reindex(columns=["Data", "Min", "Max"])

            #Tranformando os dados de string para float
            leitor["Max"] = pd.to_numeric(leitor["Max"], errors="coerce")#.astype(float)
            leitor["Min"] = pd.to_numeric(leitor["Min"], errors="coerce")#.astype(float)

            #Resumindo os extremos da estação

            linha = leitor["Max"].idxmin()
            colunas = ["Data", "Min", "Max"]
            resultado = leitor.loc[linha, colunas]

            linha2 = leitor["Min"].idxmin()
            colunas2 = ["Data", "Min", "Max"]
            resultado2 = leitor.loc[linha2, colunas2]

            linha3 = leitor["Max"].idxmax()
            colunas3 = ["Data", "Min", "Max"]
            resultado3 = leitor.loc[linha3, colunas3]
            
            linha4 = leitor["Min"].idxmax()
            colunas4 = ["Data", "Min", "Max"]
            resultado4 = leitor.loc[linha4, colunas4]

            leitor["Data"] = pd.to_datetime(leitor["Data"])

            leitor["Dia"] = leitor["Data"].dt.day
            leitor["Mes"] = leitor["Data"].dt.month
            leitor["Ano"] = leitor["Data"].dt.year

                #Imprimindo os resultados na tela
            dados1 = ("A estação de {}, registrou sua menor máxima em {}, com {} graus, e a maior em {} com {}.".format(primeiraLinha, resultado["Data"], resultado["Max"], resultado3["Data"], resultado3["Max"]))
            dados2 = ("Nas mínimas, a menor foi {} graus em {}, e a maior foi {}, em {}.".format(resultado2["Min"], resultado2["Data"], resultado4["Min"], resultado4["Data"]))

            naTela = tk.Label(janelaUsuario, text=dados1 + "\n" + dados2)
            naTela.pack()

            print(dados1, dados2)

    botaoSelecaoXLSX.config(state="normal")
    botao_selecaoCSV.config(state="disabled")

def planilhaDoUsuario():
    global planilha
    global caminho_planilha
    while True:
        caminho_planilha = filedialog.askopenfilename(initialdir=os.path.expanduser("~/Downloads"), title="Selecione a planilha CSV de destino dos dados", filetypes=(("Arquivocs XLSX", "*.xlsx"), ("Todos os arquivos", "*.*")))
        while not caminho_planilha.endswith(".xlsx"):
            if not caminho_planilha:
                return
            messagebox.showerror("Formato não aceito", "Para a correta extração dos dados, você precisa escolher uma planilha no formato Excel (.xlsx).")
            caminho_planilha = filedialog.askopenfilename(initialdir=os.path.expanduser("~/Downloads"), title="Selecione a planilha CSV de destino dos dados", filetypes=(("Arquivocs XLSX", "*.xlsx"), ("Todos os arquivos", "*.*")))
        if caminho_planilha:
            alerta = messagebox.askquestion("ATENÇÃO!", "Para que os dados sejam corretamente escritos, a planilha de destino precisa ter cada ano separado por abas, com cada aba contendo como título o ano. As temperaturas de TODOS OS ANOS precisam começar na célula B3, ou seja, na linha 3 e coluna 2. Não me responssbilizo por eventuais perdas de dados devido ao uso sem seguir esses critérios. Você pode fazer uma cópia da sua planilha e ver se ela é compatível com o formato aceito por SHEETSMASTER, ou cancelar essa operação fechando a janela. DESEJA CONTINUAR? ")
            if alerta == 'no':
                continue
            planilha = load_workbook(caminho_planilha)
            adicionandoNaPlanilha(planilha)
            return

        
def adicionandoNaPlanilha(planilha):
    botaoSelecaoXLSX.config(state="disabled")
    primeiroAno = min(int(sheet) for sheet in planilha.sheetnames if sheet.isnumeric())
    anos = leitor["Ano"].dropna().unique()

    for ano in leitor["Ano"].dropna().unique():
        if ano < primeiroAno:
             continue
        
        dadosAno = leitor[leitor["Ano"]==ano]
        print(f"Ano: {ano}")
        print(f"Dados do ano: {dadosAno}")
        organizado = pd.DataFrame()

        for mes in range(1, 13):
            dadosMes = dadosAno[dadosAno["Mes"] == mes]
            minimas = dadosMes[["Dia", "Min"]].set_index("Dia")
            maximas = dadosMes[["Dia", "Max"]].set_index("Dia")
            minimas.columns = [f"Mínima{mes}"]
            maximas.columns = [f"Máxima{mes}"]
            organizado.index = organizado.index.astype(str)
            inicio = pd.to_datetime(f"{int(ano)}-{int(mes):02d}-01")
            fim = inicio + pd.offsets.MonthEnd(0)
            dias = pd.date_range(inicio, fim)
            dias = dias.day
            minimas = minimas.reindex(dias)
            maximas = maximas.reindex(dias)
            organizado = organizado.reset_index(drop=True)
            minimas = minimas.reset_index(drop=True)
            maximas = maximas.reset_index(drop=True)
            organizado = pd.concat([organizado, minimas, maximas], axis=1)
            print(f"Organizado: {organizado}")

        sheet_name = str(int(ano))
        if sheet_name in planilha.sheetnames:
                ws = planilha[sheet_name]
        else:        
                continue
                
        barraProgresso['maximum'] = organizado.shape[1]
        for col, col_data in enumerate(organizado.values.T, start=2):
            for row, value in enumerate(col_data, start=3):
                cell_value = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col, value=value)
                barraProgresso['value'] = col - 1

                #else:
                    #print(f"Not writing value {value} to cell ({row}, {col}) because cell value is {cell_value}")

    planilha.save(caminho_planilha)




    messagebox.showinfo("Operação concluída", "Sua planilha foi preenchida!!")


def finalizar():
     if adicionandoNaPlanilha:
          pararOperacacao = messagebox.askquestion("ATENÇÃO!", "A operação de adição de dados em uma planilha já foi iniciada. Parar o processo no meio pode corromper a planilha. Deseja parar o processo mesmo assim")

          if pararOperacacao == "no":
               return
          
          sys.exit()

#Botões
botao_selecaoCSV = tk.Button(janelaUsuario, text="Selecione arquivo CSV ou descompacte a pasta", command= selecionandoCSV)
botao_selecaoCSV.pack()

botaoSelecaoXLSX = tk.Button(janelaUsuario, text="Selecione a planilha de destino dos dados.", command=planilhaDoUsuario)
botaoSelecaoXLSX.config(state="disabled")
botaoSelecaoXLSX.pack()

botaoFinalizar = tk.Button(janelaUsuario, text="Finzalizar", command=finalizar)
botaoFinalizar.pack()
janelaUsuario.mainloop()